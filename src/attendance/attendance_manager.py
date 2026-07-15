"""
=========================================================
Attendance Manager
=========================================================
"""

from common import *

from openpyxl import Workbook, load_workbook
from datetime import datetime

ATTENDANCE_FILE = PROJECT_ROOT / "attendance" / "Attendance.xlsx"

def mark_attendance(student_id, student_name):

    ATTENDANCE_FILE.parent.mkdir(parents=True, exist_ok=True)

    today = datetime.now().strftime("%Y-%m-%d")
    current_time = datetime.now().strftime("%H:%M:%S")

    # Create workbook if it doesn't exist
    if not ATTENDANCE_FILE.exists():

        workbook = Workbook()
        sheet = workbook.active

        sheet.title = "Attendance"

        sheet.append([
            "Student ID",
            "Student Name",
            "Date",
            "Time",
            "Status"
        ])

        workbook.save(ATTENDANCE_FILE)

    workbook = load_workbook(ATTENDANCE_FILE)
    sheet = workbook.active

    # Prevent duplicate attendance
    for row in sheet.iter_rows(min_row=2, values_only=True):

        if row[0] == student_id and row[2] == today:
            workbook.close()
            return False

    sheet.append([
        student_id,
        student_name,
        today,
        current_time,
        "Present"
    ])

    workbook.save(ATTENDANCE_FILE)
    workbook.close()

    return True